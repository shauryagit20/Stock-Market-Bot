import openpyxl
import os
import datetime
import yfinance
historical_data_path  =  "Over here enter the path to the directory where the historical data is saved"
saved_path = "Over here enter the path of the folder where the main.py file is save"
def Download_Dataheet():

    wb = openpyxl.load_workbook("ind_nifty500list.xlsx")
    ws = wb.active
    no_of_rows = ws.max_row
    col_name = "C"
    todays_date = datetime.datetime.today().strftime("%Y-%m-%d")
    os.chdir(historical_data_path)
    for files in os.listdir():
        os.remove(files)
    for rows in range(2, no_of_rows + 1):
        cell_addr = col_name + str(rows)
        cell_value = str(ws[cell_addr].value) + ".NS"
        data = yfinance.download(tickers=cell_value, start="2020-12-01", end=todays_date)
        os.chdir(historical_data_path)
        wb1 = openpyxl.Workbook
        ws1 = wb1.active
        filename = cell_value + ".xlsx"
        wb.save(filename)
        data.to_excel(filename)
        os.chdir(saved_path)

