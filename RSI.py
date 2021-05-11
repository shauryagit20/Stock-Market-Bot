import os
import openpyxl


class RSI:

    def __init__(self, filename):
        self.filename = filename
        self.RSI = 0
        self.rs = 0
        self.up = 0
        self.down = 0
        self.up_avg = 0
        self.down_avg = 0
        self.historical_data = []
        self.wb = openpyxl.load_workbook(filename)
        self.ws = self.wb.active
        self.col_name = "E"
        self.historical_data_path = "Over here enter the path to the directory where the historical data is saved"
        self.saved_path = "Over here enter the path of the folder where the main.py file is save"

    def get_historical_data(self):
        no_of_rows = self.ws.max_row
        last_row = no_of_rows - 14

        for rows in range(last_row, no_of_rows + 1):
            cell_addr = self.col_name + str(rows)

            cell_value = float(self.ws[cell_addr].value)
            self.historical_data.append(cell_value)


    def calc_rs(self):
        for index in range(1, len(self.historical_data)):
            prev_value = self.historical_data[index - 1]
            current_value = self.historical_data[index]
            if (current_value > prev_value):
                self.up = self.up + (current_value - prev_value)
            else:
                self.down = self.down + (prev_value - current_value)
        self.up_avg = self.up / 14
        self.down_avg = self.down / 14
        self.rs = self.up_avg / self.down_avg

    def calc_rsi(self):
        self.rs = self.rs + 1
        self.rsi = 100 - (100 / self.rs)

        print(len(self.historical_data))
        self.check_rsi()

    def check_rsi(self):
        if (self.rsi < 30):
            os.chdir(self.saved_path)
            with open("output.txt", "a") as f:
                f.write(f"{self.filename} : {self.rsi} \n " )
            os.chdir(self.historical_data_path)

    def main(self):
        self.get_historical_data()
        self.calc_rs()
        self.calc_rsi()

historical_data_path  =  "Over here enter the path to the directory where the historical data is saved"

def Execute():
    os.chdir(historical_data_path)
    list_of_files =  os.listdir()

    for files in list_of_files:
        try:
            obj1 =  RSI(filename=files)
            obj1.main()
        except:
            print("Corrupted file")

