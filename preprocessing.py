import openpyxl
import os
import time
historical_data_path  =  "Over here enter the path to the directory where the historical data is saved"
def preprocess():
    os.chdir(historical_data_path)
    list_of_files = os.listdir()
    list_of_rows = []
    for files in list_of_files:
        try:

            wb = openpyxl.load_workbook(files)
            ws = wb.active
            no_of_rows = ws.max_row
            list_of_rows.append(no_of_rows)
        except Exception as e:
            print(e)
            print("Removing the file")
            os.remove(files)
    counter = 0
    current = 0
    max_row = 0
    for ele in list_of_rows:
        current = list_of_rows.count(ele)
        if current > counter:
            max_row = ele

    print(max_row)
    list_of_files = os.listdir(historical_data_path)

    for ele in list_of_files:
        wb = openpyxl.load_workbook(files)
        ws = wb.active
        no_of_rows = ws.max_row

        if (no_of_rows != max_row):
            os.remove(ele)
